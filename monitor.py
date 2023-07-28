import pandas as pd
import glob
import os
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# 1. Загрузка данных из Excel
def load_data(filename):
    return pd.read_excel(filename)

# 2. Сравнение данных
def compare_dataframes(df1, df2):
    if df1.equals(df2) or df2.empty:
        return None, df1
    else:
        changes = df1 != df2
        changes = changes & (df1.notna() | df2.notna())
        return changes, df1

# 3. Логирование изменений и выделение их в Excel
def log_and_highlight_changes(changes, current_data, filename):
    if not os.path.exists(filename):
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            current_data.to_excel(writer, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    for i, row in enumerate(current_data.iterrows(), start=2):  # начинаем с 2 из-за названий столбцов
        for j, value in enumerate(row[1], start=1):
            cell = ws.cell(row=i, column=j)
            # Если это измененная ячейка, выделяем ее красным цветом
            if changes is not None and changes.iloc[i-2, j-1]:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    wb.save(filename)

def get_latest_files(pattern, n):
    files = sorted(glob.glob(pattern), key=lambda x: datetime.strptime(x.split("_")[0], '%Y-%m-%d'))
    return files[-n:]

def main():
    all_files = glob.glob("????-??-??_df*.xlsx")
    print(all_files)
    unique_dfs = list(set([f.split('_')[-1] for f in all_files]))
    print(unique_dfs)
    for df_id in unique_dfs:
        filenames = get_latest_files(f"*_{df_id}", 2)
        
        # if len(filenames) < 2:
        #     continue

        previous_data = load_data(filenames[0])
        current_data = load_data(filenames[1])
        
        changes, data_to_log = compare_dataframes(current_data, previous_data)

        log_filename = filenames[1].replace(".xlsx", "_log.xlsx")
        
        print(log_filename)
        log_and_highlight_changes(changes, data_to_log, log_filename)

main()
